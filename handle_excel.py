# coding=utf-8
import sys
import os
base_path = os.getcwd()
sys.path.append(base_path)
import openpyxl

class HandleExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        # 使用时需修改
        open_excel = openpyxl.load_workbook(base_path + "file")
        return open_excel

    def get_sheet_data(self, index=None):
        '''
        加载所有excel内容
        '''
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, cols):
        '''
        获取某一个单元格内容
        '''
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    def get_rows(self):
        '''
        获取总行数
        '''
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self,row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        '''写入数据'''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, cols, value)
        # 使用时需修改，与上面保持一致
        wb.save(base_path + "file")


excel_data = HandleExcel()

if __name__ == "__main__":
    handle = HandleExcel()
    # print(handle.get_cell_value(3,5))
    # print(handle.get_rows_value(3))
    handle.excel_write_data(2,14,'通过cui')